#!/usr/bin/env python3
"""
Teste para verificar formatação de números grandes
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from io import BytesIO

def test_number_formatting():
    """Testa formatação de números grandes"""
    
    print("🧪 Testando formatação de números grandes...")
    
    # Dados de teste com números grandes
    data = {
        'Region': ['West', 'Northeast', 'Southeast', 'South', 'Midwest'],
        'Total': [269943182000, 186324067000, 163171236000, 144663181000, 135800459000],
        'Média': [110270900, 78419200, 68123400, 60234500, 56500200],
        'Quantidade': [2448000, 2376000, 2395000, 2401000, 2403000],
        'Percentual': [30.00, 20.70, 18.15, 16.08, 15.07]
    }
    
    df = pd.DataFrame(data)
    print(f"📊 DataFrame criado: {len(df)} linhas")
    print(f"📋 Colunas: {df.columns.tolist()}")
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Teste Formatação"
    
    # Adicionar dados
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Aplicar formatação específica
            if isinstance(value, (int, float)) and value is not None:
                if col_idx == 5:  # Coluna Percentual
                    cell.number_format = '0.00"%"'
                elif value >= 1e9:
                    cell.number_format = '#,##0.0"B"'
                elif value >= 1e6:
                    cell.number_format = '#,##0.0"M"'
                elif value >= 1e3:
                    cell.number_format = '#,##0.0"K"'
                else:
                    cell.number_format = '#,##0.00'
    
    # Ajustar larguras
    ws.column_dimensions['A'].width = 15  # Region
    ws.column_dimensions['B'].width = 25  # Total
    ws.column_dimensions['C'].width = 20  # Média
    ws.column_dimensions['D'].width = 20  # Quantidade
    ws.column_dimensions['E'].width = 15  # Percentual
    
    # Salvar
    with open('teste_formatacao.xlsx', 'wb') as f:
        wb.save(f)
    
    print("💾 Arquivo salvo como 'teste_formatacao.xlsx'")
    print("🔍 Abra o arquivo e verifique se os números estão formatados corretamente!")
    
    # Mostrar valores esperados
    print("\n📊 Valores esperados:")
    for idx, row in df.iterrows():
        print(f"{row['Region']}: {row['Total']:,.0f} -> {row['Total']/1e6:.1f}M")

if __name__ == "__main__":
    print("🎯 TESTE DE FORMATAÇÃO DE NÚMEROS")
    print("=" * 40)
    
    test_number_formatting() 