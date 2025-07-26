import pandas as pd
from io import BytesIO
from fastapi import APIRouter, HTTPException, UploadFile, File, Query, Request, Depends, Form
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from app.security import bearer_scheme
from app.models import Report, ReportCreate
from app.services.supabase_client import supabase
from app.dependencies.auth import CurrentUser
from typing import List, Optional, Dict
from datetime import datetime
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
from slowapi import Limiter
from app.limiter import limiter, UPLOAD_LIMIT, REPORT_GENERATION_LIMIT, GENERAL_LIMIT, AUTH_LIMIT
from app.utils.crypto_utils import encrypt_file_bytes, decrypt_file_bytes, encrypt_data, decrypt_data
import numpy as np

router = APIRouter()

MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB

@router.get("/test-auth")
def test_auth():
    """Endpoint de teste para verificar se o servidor está funcionando"""
    return {"message": "Servidor funcionando!", "status": "ok"}

@router.get("/test-token")
def test_token(token: str = Query(..., description="Token para teste")):
    """Endpoint de teste para verificar se o token está funcionando"""
    try:
        import jwt
        # Decodificar sem verificar assinatura
        payload = jwt.decode(token.strip(), options={"verify_signature": False})
        return {
            "message": "Token válido!",
            "user_id": payload.get('sub'),
            "email": payload.get('email'),
            "exp": payload.get('exp')
        }
    except Exception as e:
        raise HTTPException(status_code=401, detail=f"Token inválido: {str(e)}")

@router.post("/analyze-columns-simple")
async def analyze_columns_simple(
    file: UploadFile = File(...)
):
    """Lista as colunas disponíveis no arquivo (sem autenticação)"""
    import pandas as pd
    from io import BytesIO
    
    content = await file.read()
    try:
        if file.filename.endswith('.csv'):
            df = pd.read_csv(BytesIO(content))
        elif file.filename.endswith('.xlsx'):
            df = pd.read_excel(BytesIO(content))
        else:
            raise HTTPException(status_code=400, detail="Formato de arquivo não suportado")
        
        return {
            "columns": df.columns.tolist(),
            "sample_data": df.head(3).to_dict(orient="records"),
            "total_rows": len(df)
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao analisar colunas: {str(e)}")

@router.post("/reports/analyze/columns")
async def analyze_columns(
    file: UploadFile = File(...),
    token: str = Query(None, description="Token JWT para autenticação"),
    current_user: str = CurrentUser
):
    """Lista as colunas disponíveis no arquivo"""
    import pandas as pd
    from io import BytesIO
    
    content = await file.read()
    try:
        if file.filename.endswith('.csv'):
            df = pd.read_csv(BytesIO(content))
        elif file.filename.endswith('.xlsx'):
            df = pd.read_excel(BytesIO(content))
        else:
            raise HTTPException(status_code=400, detail="Formato de arquivo não suportado")
        
        return {
            "columns": df.columns.tolist(),
            "sample_data": df.head(3).to_dict(orient="records")
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao analisar colunas: {str(e)}")

@router.get("/reports", response_model=List[Report])
def list_reports(current_user: str = CurrentUser):
    try:
        print(f"🔍 Usuário autenticado: {current_user}")
        response = supabase.table("reports").select("*").order("created_at", desc=True).execute()
        return response.data
    except Exception as e:
        print(f"❌ Erro ao buscar relatórios: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/reports/{report_id}", response_model=Report)
def get_report(report_id: int, current_user: str = CurrentUser):
    try:
        response = supabase.table("reports").select("*").eq("id", report_id).single().execute()
        return response.data
    except Exception as e:
        raise HTTPException(status_code=404, detail="Relatório não encontrado: " + str(e))

@router.post("/reports/upload")
@limiter.limit("30/minute")
async def upload_report_file(
    file: UploadFile = File(...),
    max_rows: int = Query(None, description="Número máximo de linhas a retornar no preview. Se não informado, retorna tudo."),
    current_user: str = CurrentUser,
    request: Request = None,
    credentials: HTTPAuthorizationCredentials = Depends(bearer_scheme)
):
    content = await file.read()
    if not file.filename:
        raise HTTPException(status_code=400, detail="Arquivo sem nome não suportado")
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="Arquivo excede o tamanho máximo de 5MB")
    try:
        # Criptografa o arquivo antes de processar
        encrypted_content = encrypt_file_bytes(content)
        # Para processar o DataFrame, descriptografa
        decrypted_content = decrypt_file_bytes(encrypted_content)
        if file.filename.endswith('.csv'):
            df = pd.read_csv(BytesIO(decrypted_content))
        elif file.filename.endswith('.xlsx'):
            df = pd.read_excel(BytesIO(decrypted_content))
        else:
            raise HTTPException(status_code=400, detail="Formato de arquivo não suportado")
        df = df.where(pd.notnull(df), None)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao processar arquivo: {str(e)}")
    if max_rows is not None:
        preview = df.head(max_rows).to_dict(orient="records")
    else:
        preview = df.to_dict(orient="records")
    return {
        "columns": df.columns.tolist(),
        "preview": preview
    }

@router.post("/reports", response_model=Report)
@limiter.limit("5/minute")
def create_report(report: ReportCreate, current_user: str = CurrentUser, request: Request = None, credentials: HTTPAuthorizationCredentials = Depends(bearer_scheme)):
    try:
        data = report.dict()
        data["created_at"] = datetime.utcnow().isoformat()
        # Exemplo: criptografar campo sensível se existir
        if "token" in data:
            data["token"] = encrypt_data(data["token"])
        if "cpf" in data:
            data["cpf"] = encrypt_data(data["cpf"])
        response = supabase.table("reports").insert(data).execute()
        return response.data[0]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.delete("/reports/{report_id}")
@limiter.limit("5/minute")
def delete_report(report_id: int, current_user: str = CurrentUser, request: Request = None, credentials: HTTPAuthorizationCredentials = Depends(bearer_scheme)):
    try:
        response = supabase.table("reports").delete().eq("id", report_id).execute()
        return {"message": "Relatório deletado com sucesso"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/reports/analyze")
async def analyze_report_file(
    file: UploadFile = File(...),
    current_user: str = CurrentUser
):
    import pandas as pd
    from io import BytesIO
    content = await file.read()
    try:
        if file.filename.endswith('.csv'):
            df = pd.read_csv(BytesIO(content))
        elif file.filename.endswith('.xlsx'):
            df = pd.read_excel(BytesIO(content))
        else:
            raise HTTPException(status_code=400, detail="Formato de arquivo não suportado")
        # Estatísticas descritivas
        stats = df.describe(include='all').to_dict()
        # Tipos de dados
        dtypes = df.dtypes.apply(lambda x: str(x)).to_dict()
        # Nulos por coluna
        nulls = df.isnull().sum().to_dict()
        # Sugestão de gráficos
        suggestions = []
        for col in df.columns:
            if pd.api.types.is_numeric_dtype(df[col]):
                suggestions.append({"column": col, "suggestion": "histogram/bar"})
            elif pd.api.types.is_categorical_dtype(df[col]) or df[col].dtype == object:
                suggestions.append({"column": col, "suggestion": "pie/bar"})
        return {
            "columns": list(df.columns),
            "dtypes": dtypes,
            "nulls": nulls,
            "stats": stats,
            "suggestions": suggestions
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao analisar arquivo: {str(e)}")

@router.post("/reports/analyze/custom")
async def analyze_custom_report_file(
    file: UploadFile = File(...),
    question: str = Form(...),
    current_user: str = CurrentUser
):
    import pandas as pd
    from io import BytesIO
    content = await file.read()
    try:
        if file.filename.endswith('.csv'):
            df = pd.read_csv(BytesIO(content))
        elif file.filename.endswith('.xlsx'):
            df = pd.read_excel(BytesIO(content))
        else:
            raise HTTPException(status_code=400, detail="Formato de arquivo não suportado")
        # Lógica simples para perguntas comuns
        q = question.lower()
        if "média" in q or "media" in q or "mean" in q:
            for col in df.select_dtypes(include='number').columns:
                if col in q:
                    return {"column": col, "mean": df[col].mean()}
            return {"means": df.mean(numeric_only=True).to_dict()}
        elif "soma" in q or "sum" in q:
            for col in df.select_dtypes(include='number').columns:
                if col in q:
                    return {"column": col, "sum": df[col].sum()}
            return {"sums": df.sum(numeric_only=True).to_dict()}
        elif "contar" in q or "count" in q:
            return {"count": len(df)}
        # Pronto para integração futura com IA
        return {"message": "Pergunta recebida, mas só posso responder perguntas simples de média, soma ou contagem por enquanto.", "question": question}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao analisar arquivo: {str(e)}")

@router.post("/reports/analyze/trend")
async def analyze_trend(
    file: UploadFile = File(...),
    date_col: str = Form(...),
    value_col: str = Form(...),
    freq: str = Form("M"),  # Mês por padrão
    current_user: str = CurrentUser
):
    """
    Analisa tendências e sazonalidade de uma métrica ao longo do tempo.
    """
    import pandas as pd
    from io import BytesIO
    content = await file.read()
    try:
        if file.filename.endswith('.csv'):
            df = pd.read_csv(BytesIO(content))
        elif file.filename.endswith('.xlsx'):
            df = pd.read_excel(BytesIO(content))
        else:
            raise HTTPException(status_code=400, detail="Formato de arquivo não suportado")
        df[date_col] = pd.to_datetime(df[date_col], errors="coerce")
        df = df.dropna(subset=[date_col, value_col])
        df = df.sort_values(date_col)
        ts = df.groupby(df[date_col].dt.to_period(freq))[value_col].sum()
        trend = ts.rolling(window=3, min_periods=1).mean().tolist()
        return {
            "periods": ts.index.astype(str).tolist(),
            "values": ts.tolist(),
            "trend": trend
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao analisar tendência: {str(e)}")

@router.post("/reports/analyze/risk-score")
async def analyze_risk_score(
    file: UploadFile = File(...),
    score_cols: str = Form(...),  # Ex: "idade,renda,dividas"
    weights: str = Form(None),    # Ex: "0.3,0.5,0.2"
    current_user: str = CurrentUser
):
    """
    Calcula score de risco/probabilidade baseado em variáveis do Excel.
    """
    import pandas as pd
    from io import BytesIO
    content = await file.read()
    try:
        if file.filename.endswith('.csv'):
            df = pd.read_csv(BytesIO(content))
        elif file.filename.endswith('.xlsx'):
            df = pd.read_excel(BytesIO(content))
        else:
            raise HTTPException(status_code=400, detail="Formato de arquivo não suportado")
        cols = [c.strip() for c in score_cols.split(",")]
        if weights:
            w = [float(x) for x in weights.split(",")]
            if len(w) != len(cols):
                raise HTTPException(status_code=400, detail="Número de pesos diferente do número de colunas")
        else:
            w = [1.0/len(cols)]*len(cols)
        df = df.dropna(subset=cols)
        # Normalizar colunas
        for i, c in enumerate(cols):
            col_min = df[c].min()
            col_max = df[c].max()
            if col_max > col_min:
                df[c] = (df[c] - col_min) / (col_max - col_min)
            else:
                df[c] = 0.0
        df["risk_score"] = np.dot(df[cols], w)
        return {
            "scores": df["risk_score"].round(3).tolist(),
            "summary": {
                "min": float(df["risk_score"].min()),
                "max": float(df["risk_score"].max()),
                "mean": float(df["risk_score"].mean()),
                "std": float(df["risk_score"].std())
            }
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao calcular score de risco: {str(e)}")

@router.post("/reports/analyze/geography")
async def analyze_geography(
    file: UploadFile = File(...),
    geo_col: str = Form(...),
    value_col: str = Form(None),
    agg: str = Form("count"),  # count, sum, mean
    token: str = Query(None, description="Token JWT para autenticação"),
    current_user: str = CurrentUser
):
    """
    Agrupa e sumariza dados por região, estado ou cidade.
    """
    import pandas as pd
    from io import BytesIO
    content = await file.read()
    try:
        if file.filename.endswith('.csv'):
            df = pd.read_csv(BytesIO(content))
        elif file.filename.endswith('.xlsx'):
            df = pd.read_excel(BytesIO(content))
        else:
            raise HTTPException(status_code=400, detail="Formato de arquivo não suportado")
        
        if geo_col not in df.columns:
            raise HTTPException(status_code=400, detail="Coluna geográfica não encontrada")
        
        # Configurar pandas para não usar notação científica
        pd.set_option('display.float_format', lambda x: '%.2f' % x)
        
        if agg == "count":
            result = df[geo_col].value_counts().to_dict()
            # Converter para formato adequado para gráficos
            chart_data = {
                "labels": list(result.keys()),
                "values": list(result.values()),
                "chart_type": "bar",
                "title": f"Contagem por {geo_col}",
                "subtitle": f"Total de registros: {sum(result.values())}"
            }
        elif agg in ("sum", "mean") and value_col:
            if value_col not in df.columns:
                raise HTTPException(status_code=400, detail="Coluna de valor não encontrada")
            
            if agg == "sum":
                grouped = df.groupby(geo_col)[value_col].sum()
                title = f"Soma de {value_col} por {geo_col}"
            else:
                grouped = df.groupby(geo_col)[value_col].mean()
                title = f"Média de {value_col} por {geo_col}"
            
            # Ordenar por valor decrescente
            grouped = grouped.sort_values(ascending=False)
            
            # Formatar valores para evitar notação científica
            formatted_values = []
            for val in grouped.values:
                if val >= 1e6:
                    formatted_values.append(f"{val/1e6:.2f}M")
                elif val >= 1e3:
                    formatted_values.append(f"{val/1e3:.2f}K")
                else:
                    formatted_values.append(f"{val:.2f}")
            
            chart_data = {
                "labels": list(grouped.index),
                "values": list(grouped.values),
                "formatted_values": formatted_values,
                "chart_type": "bar",
                "title": title,
                "subtitle": f"Total: {grouped.sum():,.2f}" if agg == "sum" else f"Média geral: {grouped.mean():,.2f}",
                "total": float(grouped.sum()) if agg == "sum" else float(grouped.mean()),
                "count": len(grouped)
            }
        else:
            raise HTTPException(status_code=400, detail="Agregação não suportada ou coluna de valor ausente")
        
        return {
            "geo_summary": grouped.to_dict() if 'grouped' in locals() else result,
            "chart_data": chart_data,
            "metadata": {
                "geo_column": geo_col,
                "value_column": value_col,
                "aggregation": agg,
                "total_records": len(df)
            }
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao analisar geografia: {str(e)}")

@router.get("/reports/dashboard/summary")
async def dashboard_summary(current_user: str = CurrentUser):
    try:
        response = supabase.table("reports").select("*").eq("user_id", current_user).execute()
        import pandas as pd
        df = pd.DataFrame(response.data)
        if df.empty:
            return {"message": "Nenhum dado para dashboard"}
        # Exemplo de sumarização: total, por mês, por título
        summary = {
            "total_reports": len(df),
            "created_per_month": df["created_at"].str[:7].value_counts().sort_index().to_dict(),
            "reports_by_title": df["title"].value_counts().to_dict() if "title" in df.columns else {}
        }
        return summary
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# Função utilitária para aplicar filtros em DataFrame
def apply_filters(df, start_date: Optional[str], end_date: Optional[str], filter_col: Optional[str], filter_val: Optional[str]):
    if start_date:
        df = df[df['created_at'] >= start_date]
    if end_date:
        df = df[df['created_at'] <= end_date]
    if filter_col and filter_val and filter_col in df.columns:
        df = df[df[filter_col] == filter_val]
    return df

@router.get("/reports/dashboard/column/{col}")
async def dashboard_column(
    col: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    filter_col: Optional[str] = None,
    filter_val: Optional[str] = None,
    current_user: str = CurrentUser
):
    response = supabase.table("reports").select("*").eq("user_id", current_user).execute()
    import pandas as pd
    df = pd.DataFrame(response.data)
    if col not in df.columns:
        raise HTTPException(status_code=400, detail="Coluna não encontrada")
    df = apply_filters(df, start_date, end_date, filter_col, filter_val)
    counts = df[col].value_counts().to_dict()
    return {"column": col, "distribution": counts}

@router.get("/reports/dashboard/stats/{col}")
async def dashboard_stats(
    col: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    filter_col: Optional[str] = None,
    filter_val: Optional[str] = None,
    current_user: str = CurrentUser
):
    response = supabase.table("reports").select("*").eq("user_id", current_user).execute()
    import pandas as pd
    df = pd.DataFrame(response.data)
    if col not in df.columns or not pd.api.types.is_numeric_dtype(df[col]):
        raise HTTPException(status_code=400, detail="Coluna numérica não encontrada")
    df = apply_filters(df, start_date, end_date, filter_col, filter_val)
    stats = df[col].describe().to_dict()
    return {"column": col, "stats": stats}

@router.get("/reports/dashboard/timeseries")
async def dashboard_timeseries(
    col: str = "created_at",
    freq: str = "M",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    filter_col: Optional[str] = None,
    filter_val: Optional[str] = None,
    current_user: str = CurrentUser
):
    response = supabase.table("reports").select("*").eq("user_id", current_user).execute()
    import pandas as pd
    df = pd.DataFrame(response.data)
    if col not in df.columns:
        raise HTTPException(status_code=400, detail="Coluna não encontrada")
    df = apply_filters(df, start_date, end_date, filter_col, filter_val)
    df[col] = pd.to_datetime(df[col], errors="coerce")
    ts = df.groupby(df[col].dt.to_period(freq)).size().sort_index().to_dict()
    return {"column": col, "freq": freq, "timeseries": ts}

@router.get("/reports/dashboard/pivot")
async def dashboard_pivot(
    col1: str,
    col2: str,
    agg: str = "count",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    filter_col: Optional[str] = None,
    filter_val: Optional[str] = None,
    current_user: str = CurrentUser
):
    response = supabase.table("reports").select("*").eq("user_id", current_user).execute()
    import pandas as pd
    df = pd.DataFrame(response.data)
    if col1 not in df.columns or col2 not in df.columns:
        raise HTTPException(status_code=400, detail="Colunas não encontradas")
    df = apply_filters(df, start_date, end_date, filter_col, filter_val)
    if agg == "count":
        pivot = pd.pivot_table(df, index=col1, columns=col2, aggfunc="size", fill_value=0)
    elif agg == "sum":
        num_cols = df.select_dtypes(include='number').columns
        if len(num_cols) == 0:
            raise HTTPException(status_code=400, detail="Nenhuma coluna numérica para somar")
        pivot = pd.pivot_table(df, index=col1, columns=col2, values=num_cols[0], aggfunc="sum", fill_value=0)
    else:
        raise HTTPException(status_code=400, detail="Agregação não suportada")
    return {"pivot": pivot.to_dict()}

@router.get("/reports/dashboard/outliers/{col}")
async def dashboard_outliers(
    col: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    filter_col: Optional[str] = None,
    filter_val: Optional[str] = None,
    current_user: str = CurrentUser
):
    response = supabase.table("reports").select("*").eq("user_id", current_user).execute()
    import pandas as pd
    df = pd.DataFrame(response.data)
    if col not in df.columns or not pd.api.types.is_numeric_dtype(df[col]):
        raise HTTPException(status_code=400, detail="Coluna numérica não encontrada")
    df = apply_filters(df, start_date, end_date, filter_col, filter_val)
    q1 = df[col].quantile(0.25)
    q3 = df[col].quantile(0.75)
    iqr = q3 - q1
    lower = q1 - 1.5 * iqr
    upper = q3 + 1.5 * iqr
    outliers = df[(df[col] < lower) | (df[col] > upper)][col].tolist()
    return {"column": col, "outliers": outliers, "lower": lower, "upper": upper}

@router.post("/reports/generate-excel")
@limiter.limit("10/minute")
async def generate_excel_report(
    file: UploadFile = File(...),
    report_type: str = Form(..., description="Tipo de relatório: 'summary', 'geography', 'trend', 'risk', 'custom'"),
    title: str = Form("Relatório Automático", description="Título do relatório"),
    include_charts: bool = Form(True, description="Incluir gráficos no Excel"),
    include_summary: bool = Form(True, description="Incluir resumo executivo"),
    custom_analysis: Optional[str] = Form(None, description="Análise customizada (para report_type='custom')"),
    value_column: Optional[str] = Form(None, description="Coluna de valores para análises"),
    region_column: Optional[str] = Form(None, description="Coluna de região para análise geográfica"),
    date_column: Optional[str] = Form(None, description="Coluna de data para análise temporal"),
    current_user: str = CurrentUser,
    request: Request = None
):
    """
    Gera relatório em Excel automaticamente com base nos dados fornecidos
    """
    try:
        # Ler o arquivo
        content = await file.read()
        if file.filename.endswith('.csv'):
            df = pd.read_csv(BytesIO(content))
        elif file.filename.endswith('.xlsx'):
            df = pd.read_excel(BytesIO(content))
        else:
            raise HTTPException(status_code=400, detail="Formato de arquivo não suportado")
        
        # Criar arquivo Excel na memória
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Configurar pandas para não usar notação científica
            pd.set_option('display.float_format', lambda x: '%.2f' % x)
            
            # Aba 1: Dados Originais
            df.to_excel(writer, sheet_name='Dados Originais', index=False)
            
            # Aba 2: Resumo Estatístico
            if include_summary:
                summary_stats = df.describe()
                summary_stats.to_excel(writer, sheet_name='Resumo Estatístico')
                
                # Adicionar informações gerais
                summary_info = pd.DataFrame({
                    'Métrica': ['Total de Registros', 'Total de Colunas', 'Data de Geração'],
                    'Valor': [len(df), len(df.columns), datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
                })
                summary_info.to_excel(writer, sheet_name='Resumo Estatístico', startrow=len(summary_stats) + 3, index=False)
            
            # Aba 3: Análise Específica baseada no tipo
            print(f"🔍 Debug - report_type: '{report_type}'")
            print(f"🔍 Debug - region_column: '{region_column}'")
            print(f"🔍 Debug - value_column: '{value_column}'")
            print(f"🔍 Debug - include_charts: {include_charts}")
            print(f"🔍 Debug - Colunas disponíveis: {df.columns.tolist()}")
            
            if report_type == 'summary':
                # Análise geral
                analysis_sheet = pd.DataFrame({
                    'Análise': ['Tipo de Relatório', 'Total de Registros', 'Colunas Disponíveis'],
                    'Valor': ['Resumo Geral', len(df), ', '.join(df.columns.tolist())]
                })
                analysis_sheet.to_excel(writer, sheet_name='Análise Geral', index=False)
                
            elif report_type == 'geography' and region_column and value_column and region_column.strip() and value_column.strip():
                print(f"✅ Condição geography atendida!")
                # Análise geográfica
                if region_column in df.columns and value_column in df.columns:
                    print(f"✅ Colunas encontradas no DataFrame!")
                    geo_analysis = df.groupby(region_column)[value_column].agg(['sum', 'mean', 'count']).reset_index()
                    geo_analysis.columns = [region_column, 'Total', 'Média', 'Quantidade']
                    geo_analysis['Percentual'] = (geo_analysis['Total'] / geo_analysis['Total'].sum() * 100).round(2)
                    
                    # Ordenar por total decrescente
                    geo_analysis = geo_analysis.sort_values('Total', ascending=False)
                    
                    geo_analysis.to_excel(writer, sheet_name='Análise Geográfica', index=False)
                    
                    # Formatação será aplicada na etapa final
                    print(f"📊 Dados da análise geográfica salvos - formatação será aplicada na etapa final")
                    
                    # Adicionar gráfico se include_charts for True
                    if include_charts:
                        print(f"🎯 Tentando adicionar gráfico...")
                        print(f"📊 Dados: {len(geo_analysis)} linhas")
                        print(f"📋 Colunas: {geo_analysis.columns.tolist()}")
                        
                        try:
                            from openpyxl.chart import BarChart, Reference
                            from openpyxl.chart.label import DataLabelList
                            print("✅ Módulos importados com sucesso")
                            
                            # Obter a planilha - tentar diferentes abordagens
                            try:
                                worksheet = writer.sheets['Análise Geográfica']
                                print(f"📄 Planilha obtida via writer.sheets: {worksheet.title}")
                            except:
                                # Tentar obter via workbook
                                workbook = writer.book
                                worksheet = workbook['Análise Geográfica']
                                print(f"📄 Planilha obtida via workbook: {worksheet.title}")
                            
                            # Criar gráfico de barras
                            chart = BarChart()
                            chart.title = f"Análise de {value_column} por {region_column}"
                            chart.style = 10
                            chart.x_axis.title = region_column
                            chart.y_axis.title = value_column
                            print("📈 Gráfico criado")
                            
                            # Dados para o gráfico (Total por região)
                            data = Reference(worksheet, min_col=2, min_row=1, max_row=len(geo_analysis)+1, max_col=2)
                            cats = Reference(worksheet, min_col=1, min_row=2, max_row=len(geo_analysis)+1)
                            print(f"📊 Referências criadas: dados={data}, categorias={cats}")
                            
                            chart.add_data(data, titles_from_data=True)
                            chart.set_categories(cats)
                            print("📊 Dados adicionados ao gráfico")
                            
                            # Adicionar rótulos de dados
                            chart.dataLabels = DataLabelList()
                            chart.dataLabels.showVal = True
                            print("🏷️ Rótulos adicionados")
                            
                            # Inserir gráfico na planilha
                            worksheet.add_chart(chart, "G2")
                            print(f"✅ Gráfico adicionado com sucesso na posição G2")
                            
                        except Exception as e:
                            print(f"❌ Erro ao adicionar gráfico: {str(e)}")
                            import traceback
                            print(f"🔍 Traceback: {traceback.format_exc()}")
                            # Se openpyxl.chart não estiver disponível, continuar sem gráfico
                            pass
                else:
                    print(f"❌ Colunas não encontradas: {region_column} ou {value_column}")
                    print(f"📋 Colunas disponíveis: {df.columns.tolist()}")
                
            elif report_type == 'trend' and date_column and value_column:
                # Análise temporal
                if date_column in df.columns and value_column in df.columns:
                    # Converter para datetime se possível
                    try:
                        df[date_column] = pd.to_datetime(df[date_column])
                        trend_analysis = df.groupby(df[date_column].dt.to_period('M'))[value_column].agg(['sum', 'mean']).reset_index()
                        trend_analysis.columns = ['Período', 'Total', 'Média']
                        trend_analysis.to_excel(writer, sheet_name='Análise Temporal', index=False)
                    except:
                        # Se não conseguir converter, agrupar por valores únicos
                        trend_analysis = df.groupby(date_column)[value_column].agg(['sum', 'mean']).reset_index()
                        trend_analysis.columns = [date_column, 'Total', 'Média']
                        trend_analysis.to_excel(writer, sheet_name='Análise Temporal', index=False)
                
            elif report_type == 'risk' and value_column:
                # Análise de risco
                if value_column in df.columns:
                    # Calcular estatísticas de risco
                    risk_stats = pd.DataFrame({
                        'Métrica': ['Média', 'Desvio Padrão', 'Mínimo', 'Máximo', 'Mediana'],
                        'Valor': [
                            df[value_column].mean(),
                            df[value_column].std(),
                            df[value_column].min(),
                            df[value_column].max(),
                            df[value_column].median()
                        ]
                    })
                    risk_stats.to_excel(writer, sheet_name='Análise de Risco', index=False)
                    
                    # Identificar outliers
                    Q1 = df[value_column].quantile(0.25)
                    Q3 = df[value_column].quantile(0.75)
                    IQR = Q3 - Q1
                    outliers = df[(df[value_column] < Q1 - 1.5 * IQR) | (df[value_column] > Q3 + 1.5 * IQR)]
                    if len(outliers) > 0:
                        outliers.to_excel(writer, sheet_name='Outliers', index=False)
                
            elif report_type == 'custom' and custom_analysis:
                # Análise customizada
                custom_sheet = pd.DataFrame({
                    'Análise Customizada': [custom_analysis],
                    'Data de Geração': [datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                    'Total de Registros': [len(df)]
                })
                custom_sheet.to_excel(writer, sheet_name='Análise Customizada', index=False)
            
            # Aba 4: Metadados
            metadata = pd.DataFrame({
                'Campo': ['Título do Relatório', 'Tipo de Análise', 'Arquivo Original', 'Data de Geração', 'Usuário'],
                'Valor': [title, report_type, file.filename, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), current_user]
            })
            metadata.to_excel(writer, sheet_name='Metadados', index=False)
        
        # Aplicar formatação aos números
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Salvar temporariamente e recarregar para aplicar formatação
            temp_output = BytesIO()
            temp_output.write(output.getvalue())
            temp_output.seek(0)
            
            wb = load_workbook(temp_output)
            
            # Aplicar formatação a todas as planilhas
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                print(f"🔧 Formatando planilha: {sheet_name}")
                
                # Formatar cabeçalhos
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                
                # Formatação específica para Análise Geográfica
                if sheet_name == 'Análise Geográfica':
                    print(f"🎯 Aplicando formatação específica para Análise Geográfica")
                    
                    # Definir larguras específicas
                    ws.column_dimensions['A'].width = 15  # Region
                    ws.column_dimensions['B'].width = 25  # Total
                    ws.column_dimensions['C'].width = 20  # Média
                    ws.column_dimensions['D'].width = 20  # Quantidade
                    ws.column_dimensions['E'].width = 15  # Percentual
                    
                    # Formatar cada coluna especificamente
                    for row in range(2, ws.max_row + 1):
                        # Coluna B - Total (formatação de milhões)
                        cell_b = ws[f'B{row}']
                        if isinstance(cell_b.value, (int, float)) and cell_b.value is not None:
                            if cell_b.value >= 1e9:
                                cell_b.number_format = '#,##0.0"B"'
                            elif cell_b.value >= 1e6:
                                cell_b.number_format = '#,##0.0"M"'
                            elif cell_b.value >= 1e3:
                                cell_b.number_format = '#,##0.0"K"'
                            else:
                                cell_b.number_format = '#,##0.00'
                            print(f"📊 Célula B{row}: {cell_b.value} -> formatada")
                        
                        # Coluna C - Média (formatação de milhares)
                        cell_c = ws[f'C{row}']
                        if isinstance(cell_c.value, (int, float)) and cell_c.value is not None:
                            if cell_c.value >= 1e6:
                                cell_c.number_format = '#,##0.0"M"'
                            elif cell_c.value >= 1e3:
                                cell_c.number_format = '#,##0.0"K"'
                            else:
                                cell_c.number_format = '#,##0.00'
                        
                        # Coluna D - Quantidade (formatação de milhares)
                        cell_d = ws[f'D{row}']
                        if isinstance(cell_d.value, (int, float)) and cell_d.value is not None:
                            if cell_d.value >= 1e6:
                                cell_d.number_format = '#,##0.0"M"'
                            elif cell_d.value >= 1e3:
                                cell_d.number_format = '#,##0.0"K"'
                            else:
                                cell_d.number_format = '#,##0.00'
                        
                        # Coluna E - Percentual
                        cell_e = ws[f'E{row}']
                        if isinstance(cell_e.value, (int, float)) and cell_e.value is not None:
                            cell_e.number_format = '0.00"%"'
                
                else:
                    # Formatação padrão para outras planilhas
                    for row in ws.iter_rows(min_row=2):
                        for cell in row:
                            if isinstance(cell.value, (int, float)) and cell.value is not None:
                                if cell.value >= 1e9:
                                    cell.number_format = '#,##0.0"B"'
                                elif cell.value >= 1e6:
                                    cell.number_format = '#,##0.0"M"'
                                elif cell.value >= 1e3:
                                    cell.number_format = '#,##0.0"K"'
                                else:
                                    cell.number_format = '#,##0.00'
                    
                    # Ajustar largura das colunas para outras planilhas
                    for column in ws.columns:
                        max_length = 0
                        column_letter = get_column_letter(column[0].column)
                        for cell in column:
                            try:
                                cell_value = str(cell.value) if cell.value is not None else ""
                                if len(cell_value) > max_length:
                                    max_length = len(cell_value)
                            except:
                                pass
                        adjusted_width = min(max_length + 4, 60)
                        ws.column_dimensions[column_letter].width = adjusted_width
            
            print(f"✅ Formatação aplicada com sucesso!")
            
            # Salvar com formatação
            final_output = BytesIO()
            wb.save(final_output)
            final_output.seek(0)
            
            filename = f"relatorio_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            return StreamingResponse(
                final_output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
            
        except Exception as e:
            print(f"❌ Erro na formatação: {str(e)}")
            import traceback
            print(f"🔍 Traceback: {traceback.format_exc()}")
            # Se a formatação falhar, retornar sem formatação
            output.seek(0)
            filename = f"relatorio_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            return StreamingResponse(
                BytesIO(output.getvalue()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao gerar relatório Excel: {str(e)}")

@router.post("/reports/generate-excel-from-template")
@limiter.limit("10/minute")
async def generate_excel_from_template(
    file: UploadFile = File(...),
    template_id: str = Form(..., description="ID do template a ser usado (ex: geography_sales, trend_monthly)"),
    title: str = Form("Relatório com Template", description="Título do relatório"),
    include_charts: bool = Form(True, description="Incluir gráficos no Excel"),
    current_user: str = CurrentUser,
    request: Request = None
):
    """
    Gera relatório em Excel usando um template pré-configurado com formatação e gráficos
    """
    try:
        # Importar templates pré-configurados
        from .templates import PREDEFINED_TEMPLATES
        
        # Buscar o template pré-configurado
        if template_id not in PREDEFINED_TEMPLATES:
            raise HTTPException(status_code=404, detail=f"Template '{template_id}' não encontrado. Templates disponíveis: {list(PREDEFINED_TEMPLATES.keys())}")
        
        template_data = PREDEFINED_TEMPLATES[template_id]
        config = template_data['config']
        
        print(f"🎯 Debug - template_id: '{template_id}'")
        print(f"🎯 Debug - analysis_type: '{config.get('analysis_type')}'")
        print(f"🎯 Debug - include_charts: {include_charts}")
        
        # Ler o arquivo
        content = await file.read()
        if file.filename.endswith('.csv'):
            df = pd.read_csv(BytesIO(content))
        elif file.filename.endswith('.xlsx'):
            df = pd.read_excel(BytesIO(content))
        else:
            raise HTTPException(status_code=400, detail="Formato de arquivo não suportado")
        
        print(f"📋 Colunas disponíveis: {df.columns.tolist()}")
        
        # Configurar pandas para não usar notação científica
        pd.set_option('display.float_format', lambda x: '%.2f' % x)
        
        # Criar arquivo Excel na memória
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Aba 1: Dados Originais
            df.to_excel(writer, sheet_name='Dados Originais', index=False)
            
            # Aba 2: Análise baseada no template
            analysis_type = config.get('analysis_type')
            
            if analysis_type == 'geography':
                value_col = config.get('value_column')
                region_col = config.get('region_column')
                
                print(f"🎯 Debug - region_column: '{region_col}'")
                print(f"🎯 Debug - value_column: '{value_col}'")
                
                if value_col in df.columns and region_col in df.columns:
                    print(f"✅ Condição geography atendida!")
                    print(f"✅ Colunas encontradas no DataFrame!")
                    
                    geo_analysis = df.groupby(region_col)[value_col].agg(['sum', 'mean', 'count']).reset_index()
                    geo_analysis.columns = [region_col, 'Total', 'Média', 'Quantidade']
                    geo_analysis['Percentual'] = (geo_analysis['Total'] / geo_analysis['Total'].sum() * 100).round(2)
                    geo_analysis.to_excel(writer, sheet_name='Análise Geográfica', index=False)
                    
                    print(f"📊 Dados da análise geográfica salvos - formatação será aplicada na etapa final")
                    
                    # Adicionar gráfico se include_charts for True
                    if include_charts:
                        print(f"🎯 Tentando adicionar gráfico...")
                        print(f"📊 Dados: {len(geo_analysis)} linhas")
                        print(f"📋 Colunas: {geo_analysis.columns.tolist()}")
                        
                        try:
                            from openpyxl.chart import BarChart, Reference
                            from openpyxl.chart.label import DataLabelList
                            print("✅ Módulos importados com sucesso")
                            
                            # Obter a planilha
                            try:
                                worksheet = writer.sheets['Análise Geográfica']
                                print(f"📄 Planilha obtida via writer.sheets: {worksheet.title}")
                            except:
                                # Tentar obter via workbook
                                workbook = writer.book
                                worksheet = workbook['Análise Geográfica']
                                print(f"📄 Planilha obtida via workbook: {worksheet.title}")
                            
                            # Criar gráfico de barras
                            chart = BarChart()
                            chart.title = f"Análise de {value_col} por {region_col}"
                            chart.style = 10
                            chart.x_axis.title = region_col
                            chart.y_axis.title = value_col
                            print("📈 Gráfico criado")
                            
                            # Dados para o gráfico (Total por região)
                            data = Reference(worksheet, min_col=2, min_row=1, max_row=len(geo_analysis)+1, max_col=2)
                            cats = Reference(worksheet, min_col=1, min_row=2, max_row=len(geo_analysis)+1)
                            print(f"📊 Referências criadas: dados={data}, categorias={cats}")
                            
                            chart.add_data(data, titles_from_data=True)
                            chart.set_categories(cats)
                            print("📊 Dados adicionados ao gráfico")
                            
                            # Adicionar rótulos de dados
                            chart.dataLabels = DataLabelList()
                            chart.dataLabels.showVal = True
                            print("🏷️ Rótulos adicionados")
                            
                            # Inserir gráfico na planilha
                            worksheet.add_chart(chart, "G2")
                            print(f"✅ Gráfico adicionado com sucesso na posição G2")
                            
                        except Exception as e:
                            print(f"❌ Erro ao adicionar gráfico: {str(e)}")
                            import traceback
                            print(f"🔍 Traceback: {traceback.format_exc()}")
                            # Se openpyxl.chart não estiver disponível, continuar sem gráfico
                            pass
                else:
                    print(f"❌ Colunas não encontradas: {region_col} ou {value_col}")
                    print(f"📋 Colunas disponíveis: {df.columns.tolist()}")
            
            elif analysis_type == 'trend':
                date_col = config.get('date_column')
                value_col = config.get('value_column')
                
                if date_col in df.columns and value_col in df.columns:
                    try:
                        df[date_col] = pd.to_datetime(df[date_col])
                        trend_analysis = df.groupby(df[date_col].dt.to_period('M'))[value_col].agg(['sum', 'mean']).reset_index()
                        trend_analysis.columns = ['Período', 'Total', 'Média']
                        trend_analysis.to_excel(writer, sheet_name='Análise Temporal', index=False)
                    except:
                        trend_analysis = df.groupby(date_col)[value_col].agg(['sum', 'mean']).reset_index()
                        trend_analysis.columns = [date_col, 'Total', 'Média']
                        trend_analysis.to_excel(writer, sheet_name='Análise Temporal', index=False)
            
            elif analysis_type == 'risk-score':
                # Análise de risco baseada no template
                risk_factors = config.get('risk_factors', [])
                if risk_factors:
                    risk_data = []
                    for factor in risk_factors:
                        col = factor.get('column')
                        if col in df.columns:
                            risk_data.append({
                                'Fator': col,
                                'Média': df[col].mean(),
                                'Desvio Padrão': df[col].std(),
                                'Peso': factor.get('weight', 0)
                            })
                    
                    if risk_data:
                        risk_df = pd.DataFrame(risk_data)
                        risk_df.to_excel(writer, sheet_name='Análise de Risco', index=False)
            
            # Aba 3: Configuração do Template
            template_info = pd.DataFrame({
                'Campo': ['Nome do Template', 'Descrição', 'Tipo de Análise', 'Configuração'],
                'Valor': [template_data['name'], template_data['description'], analysis_type, str(config)]
            })
            template_info.to_excel(writer, sheet_name='Template Info', index=False)
            
            # Aba 4: Metadados
            metadata = pd.DataFrame({
                'Campo': ['Título do Relatório', 'Template Usado', 'Arquivo Original', 'Data de Geração', 'Usuário'],
                'Valor': [title, template_data['name'], file.filename, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), current_user]
            })
            metadata.to_excel(writer, sheet_name='Metadados', index=False)
        
        # Aplicar formatação profissional
        try:
            print(f"🎨 Aplicando formatação profissional...")
            
            # Carregar o workbook para formatação
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Salvar temporariamente e recarregar para formatação
            temp_output = BytesIO()
            temp_output.write(output.getvalue())
            temp_output.seek(0)
            
            wb = load_workbook(temp_output)
            
            # Formatar cada planilha
            for sheet_name in wb.sheetnames:
                print(f"Formatando planilha: {sheet_name}")
                ws = wb[sheet_name]
                
                # Formatar cabeçalhos
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Ajustar largura das colunas
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            cell_value = str(cell.value) if cell.value is not None else ""
                            if len(cell_value) > max_length:
                                max_length = len(cell_value)
                        except:
                            pass
                    adjusted_width = min(max_length + 4, 60)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Formatação específica para Análise Geográfica
                if sheet_name == 'Análise Geográfica':
                    print(f"Aplicando formatação específica para Análise Geográfica")
                    
                    # Formatar coluna Total (coluna B) com largura extra e formatação específica
                    ws.column_dimensions['B'].width = 25
                    
                    # Formatar valores da coluna Total
                    for row in range(2, ws.max_row + 1):
                        cell = ws[f'B{row}']
                        if isinstance(cell.value, (int, float)) and cell.value is not None:
                            if cell.value >= 1e9:
                                cell.number_format = '#,##0.0"B"'
                            elif cell.value >= 1e6:
                                cell.number_format = '#,##0.0"M"'
                            elif cell.value >= 1e3:
                                cell.number_format = '#,##0.0"K"'
                            else:
                                cell.number_format = '#,##0.00'
                            print(f"Célula B{row}: {cell.value} -> formatada")
                    
                    # Formatar coluna Percentual (coluna E) como porcentagem
                    for row in range(2, ws.max_row + 1):
                        cell = ws[f'E{row}']
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '0.00"%"'
                    
                    print(f"✅ Formatação específica aplicada à planilha 'Análise Geográfica'")
                
                else:
                    # Formatação padrão para outras planilhas
                    for row in ws.iter_rows(min_row=2):
                        for cell in row:
                            if isinstance(cell.value, (int, float)) and cell.value is not None:
                                if cell.value >= 1e9:
                                    cell.number_format = '#,##0.0"B"'
                                elif cell.value >= 1e6:
                                    cell.number_format = '#,##0.0"M"'
                                elif cell.value >= 1e3:
                                    cell.number_format = '#,##0.0"K"'
                                else:
                                    cell.number_format = '#,##0.00'
            
            print(f"✅ Formatação aplicada com sucesso!")
            
            # Salvar com formatação
            final_output = BytesIO()
            wb.save(final_output)
            final_output.seek(0)
            
            filename = f"relatorio_template_{template_data['name'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            return StreamingResponse(
                final_output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
            
        except Exception as e:
            print(f"❌ Erro na formatação: {str(e)}")
            import traceback
            print(f"🔍 Traceback: {traceback.format_exc()}")
            # Se a formatação falhar, retornar sem formatação
            output.seek(0)
            filename = f"relatorio_template_{template_data['name'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            return StreamingResponse(
                BytesIO(output.getvalue()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao gerar relatório Excel com template: {str(e)}") 