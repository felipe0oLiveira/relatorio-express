#!/usr/bin/env python3
"""
Teste simples para criar Excel com gráfico usando openpyxl
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from io import BytesIO

def test_create_excel_with_chart():
    """Testa criação de Excel com gráfico"""
    
    print("🧪 Testando criação de Excel com gráfico...")
    
    # Dados de teste
    data = {
        'Region': ['West', 'Northeast', 'Southeast', 'South', 'Midwest'],
        'Total': [270000000, 186000000, 163000000, 145000000, 136000000]
    }
    
    df = pd.DataFrame(data)
    print(f"📊 DataFrame criado: {len(df)} linhas")
    print(f"📋 Colunas: {df.columns.tolist()}")
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Análise Geográfica"
    
    # Adicionar dados
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    print("📄 Dados adicionados à planilha")
    
    # Criar gráfico
    try:
        chart = BarChart()
        chart.title = "Análise de Total por Region"
        chart.style = 10
        chart.x_axis.title = "Region"
        chart.y_axis.title = "Total"
        
        # Referências
        data = Reference(ws, min_col=2, min_row=1, max_row=len(df)+1, max_col=2)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(df)+1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        # Rótulos
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        
        # Adicionar à planilha
        ws.add_chart(chart, "D2")
        
        print("✅ Gráfico criado e adicionado!")
        
        # Salvar
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Salvar arquivo
        with open('teste_grafico_simples.xlsx', 'wb') as f:
            f.write(output.getvalue())
        
        print("💾 Arquivo salvo como 'teste_grafico_simples.xlsx'")
        print("🔍 Abra o arquivo e verifique se há gráfico!")
        
        return True
        
    except Exception as e:
        print(f"❌ Erro: {str(e)}")
        import traceback
        print(f"🔍 Traceback: {traceback.format_exc()}")
        return False

if __name__ == "__main__":
    print("🎯 TESTE SIMPLES DE GRÁFICO")
    print("=" * 40)
    
    test_create_excel_with_chart() 